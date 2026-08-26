#!/usr/bin/env python3
"""把测试用例 Markdown 表格转换为禅道可导入的 CSV 或 Excel。

用法:
    python generate_excel.py <input.md> <output.csv|output.xlsx>

- 输出 .csv: 仅用标准库，写 UTF-8 BOM，中文在 Excel/禅道中不乱码。
- 输出 .xlsx: 需要 openpyxl (pip install openpyxl)，缺失时给出提示。

输入 Markdown 需包含一个管道表格 (| ... |)，表头列顺序即导出列顺序，
推荐: 所属模块 | 用例标题 | 前置条件 | 操作步骤 | 预期结果 | 优先级 | 用例类型 | 标签
单元格内多步骤用 <br> 分隔，会转换为换行。
"""
import argparse
import csv
import os
import sys


def parse_markdown_table(text):
    """从 Markdown 文本中解析第一个管道表格，返回 (headers, rows)。"""
    rows = []
    for line in text.splitlines():
        s = line.strip()
        if not (s.startswith("|") and s.endswith("|")):
            continue
        cells = [c.strip() for c in s[1:-1].split("|")]
        # 跳过分隔行，如 |---|---|
        if cells and all(set(c) <= set("-: ") and c for c in cells):
            continue
        cells = [c.replace("<br>", "\n").replace("<br/>", "\n") for c in cells]
        rows.append(cells)
    if not rows:
        raise ValueError("未在输入中找到 Markdown 表格（以 | 开头和结尾的行）。")
    headers, data = rows[0], rows[1:]
    width = len(headers)
    # 规整每行列数
    data = [(r + [""] * width)[:width] for r in data]
    return headers, data


def write_csv(path, headers, rows):
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(headers)
        w.writerows(rows)


def write_xlsx(path, headers, rows):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
    except ImportError:
        sys.exit(
            "缺少 openpyxl。请先 `pip install openpyxl`，"
            "或改用 .csv 输出（禅道同样支持 CSV 导入）。"
        )
    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in rows:
        ws.append(r)
    wrap = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows():
        for c in row:
            c.alignment = wrap
    wb.save(path)


def main():
    ap = argparse.ArgumentParser(description="测试用例 Markdown 表格 -> 禅道 CSV/Excel")
    ap.add_argument("input", help="输入 Markdown 文件")
    ap.add_argument("output", help="输出文件 (.csv 或 .xlsx)")
    args = ap.parse_args()

    if not os.path.isfile(args.input):
        sys.exit(f"输入文件不存在: {args.input}")
    with open(args.input, encoding="utf-8") as f:
        headers, rows = parse_markdown_table(f.read())

    ext = os.path.splitext(args.output)[1].lower()
    if ext == ".csv":
        write_csv(args.output, headers, rows)
    elif ext in (".xlsx", ".xlsm"):
        write_xlsx(args.output, headers, rows)
    else:
        sys.exit(f"不支持的输出扩展名: {ext}（请用 .csv 或 .xlsx）")

    print(f"已生成 {args.output}（{len(rows)} 条用例，{len(headers)} 列）")


if __name__ == "__main__":
    main()
